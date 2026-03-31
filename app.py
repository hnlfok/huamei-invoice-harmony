#!/usr/bin/env python3
"""
华美物流发货单生成器 - Flask Web版 + PWA支持
运行: python3 app.py
访问: http://localhost:5000
"""
import os, re, io
from flask import Flask, request, send_file, render_template_string, Response
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
UPLOAD = '/tmp/huamei_pwa_uploads'
os.makedirs(UPLOAD, exist_ok=True)

TEMPLATE_SHEET = "重庆华美物流有限公司重庆有研专用发货单"
ITEMS_START_ROW, ITEMS_END_ROW = 7, 14

def build_sheet(ws, recipient_name, recipient_phone, company_name,
                address, items, waybill_num, ship_date):
    ws.row_dimensions[1].height = 30
    ws["A1"] = TEMPLATE_SHEET
    ws["A1"].font = Font(name="黑体", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:G1")
    ws["B2"] = "发货日期:"; ws["C2"] = ship_date
    ws["C2"].number_format = "yyyy-mm-dd"
    ws["E2"] = "运单号："; ws["F2"] = waybill_num
    for c in ["B2","C2","E2","F2"]:
        ws[c].font = Font(name="宋体", size=11)
        ws[c].alignment = Alignment(horizontal="left", vertical="center")
    ws["A4"] = "收货人(电话):"; ws["B4"] = f"{recipient_name}{recipient_phone}"
    ws["E4"] = "收货单位:"; ws["F4"] = company_name
    ws.merge_cells("B4:C4"); ws.merge_cells("E4:G4")
    ws["A5"] = "收货地址："; ws["B5"] = address
    ws.merge_cells("B5:G5")
    headers = ["品名","规格","件数","重量（kg）","批号","是否送货","托盘"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=col, value=h)
        c.font = Font(name="宋体", size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    for i, item in enumerate(items):
        row = ITEMS_START_ROW + i
        for col, key in enumerate(["品名","规格","件数","重量","批号","是否送货","托盘"], 1):
            c = ws.cell(row=row, column=col, value=item.get(key,""))
            c.font = Font(name="宋体", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["H6"] = "第一联白联存底"; ws.merge_cells("H6:H10")
    ws["H12"] = "第二联红联回单"; ws.merge_cells("H12:H17")
    ws["H19"] = "第三联黄联收货人"; ws.merge_cells("H19:H23")
    for c in ["H6","H12","H19"]:
        ws[c].font = Font(name="宋体", size=9)
        ws[c].alignment = Alignment(horizontal="center", vertical="center")
    ws["A15"] = "合计:"
    r1,r2 = ITEMS_START_ROW, ITEMS_END_ROW
    ws["C15"] = f"=COUNTA(C{r1}:C{r2})"
    ws["D15"] = f"=SUM(D{r1}:D{r2})"
    ws["G15"] = f"=COUNTA(G{r1}:G{r2})"
    ws["A16"] = ("备注:(本发货单一式三联,第二联回单联为本公司收货及财务结算凭证,"
                 "请妥善保存。1吨及1吨以上均要送货。1吨以下有标注送货的要送货)")
    ws["A16"].font = Font(name="宋体", size=8)
    ws["A16"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A16:G18")
    ws["A20"] = "收货人签章:"; ws["E24"] = "收货日期:"
    ws["A25"] = ("重庆华美物流有限公司        联系人:杨卉梅"
                 "       电话:13667625772"
                 "       地址:重庆两江新区华荣货运市场A422")
    ws["A25"].font = Font(name="宋体", size=9)
    ws.merge_cells("A25:G25")
    for col, w in {1:22,2:10,3:14,4:10,5:12,6:10,7:10,8:16}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def gen_invoice(inp_file, ship_date):
    wb_in = openpyxl.load_workbook(inp_file)
    ws_in = wb_in.active
    rows = list(ws_in.iter_rows(values_only=True))
    header = rows[1]
    col_map = {h:i for i,h in enumerate(header) if h}
    def g(row, name):
        idx = col_map.get(name)
        return row[idx] if idx is not None else None
    groups = {}
    for row in rows[2:]:
        if not any(row): continue
        name = g(row,"收货人") or ""
        addr = g(row,"收货地址") or ""
        key = (name, addr)
        if key not in groups: groups[key] = []
        groups[key].append({
            "品名": g(row,"品名") or "", "规格": g(row,"规格") or "",
            "件数": g(row,"件数") or "", "重量": g(row,"重量") or 0,
            "批号": g(row,"批号及相应桶数") or "",
            "是否送货": g(row,"低于1吨的是否送货") or "",
            "托盘": g(row,"是否运输公司打托盘") or "",
        })
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    base = int(ship_date.replace("-","").replace("/","")) * 10000
    for si, ((name,addr), items) in enumerate(groups):
        if not items: continue
        items = [it for it in items if it["品名"]]
        pm = re.search(r"1[3-9]\d{9}", str(name))
        rname = name[:pm.start()] if pm else name
        rphone = pm.group() if pm else ""
        company = ""
        sname = f"{rname[:8]}{rphone[-11:] if rphone else ''}"
        ws = wb_out.create_sheet(sname)
        build_sheet(ws, rname, rphone, company, addr, items, base+si+1, ship_date)
    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf

# --- PWA 资源 ---
MANIFEST = {
  "name": "华美物流发货单生成器",
  "short_name": "发货单",
  "description": "输入今日发货数据，自动生成华美物流发货单",
  "start_url": "/",
  "display": "standalone",
  "background_color": "#ffffff",
  "theme_color": "#4472C4",
  "orientation": "portrait",
  "icons": [
    {"src": "/icon-192.png", "sizes": "192x192", "type": "image/png"},
    {"src": "/icon-512.png", "sizes": "512x512", "type": "image/png"}
  ]
}

SW_JS = """
const CACHE = 'huamei-pwa-v1';
self.addEventListener('install', e => {
  e.waitUntil(caches.open(CACHE).then(c => c.addAll(['/'])));
  self.skipWaiting();
});
self.addEventListener('activate', e => {
  e.waitUntil(caches.keys().then(ns => Promise.all(ns.filter(n=>n!==CACHE).map(n=>caches.delete(n)))));
  self.clients.claim();
});
self.addEventListener('fetch', e => {
  e.respondWith(caches.match(e.request).then(r => r || fetch(e.request)));
});
"""

# --- Web UI ---
HTML = """
<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="theme-color" content="#4472C4">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="default">
<title>华美物流发货单生成器</title>
<link rel="manifest" href="/manifest.json">
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:黑体;background:#f0f2f5;min-height:100vh}
  .header{background:#4472C4;color:#fff;padding:20px;text-align:center;font-size:20px}
  .card{background:#fff;border-radius:12px;padding:24px;max-width:520px;margin:20px auto;box-shadow:0 2px 12px rgba(0,0,0,.1)}
  label{display:block;margin-top:16px;font-weight:bold;color:#444;font-size:14px}
  input[type=text],input[type=file]{width:100%;padding:12px;margin-top:6px;border:1px solid #ddd;border-radius:8px;font-size:14px;background:#fff}
  button{width:100%;padding:14px;margin-top:20px;background:#4472C4;color:#fff;border:none;border-radius:8px;font-size:16px;cursor:pointer}
  button:hover{background:#2851A3}
  button:disabled{background:#ccc;cursor:not-allowed}
  .tip{background:#fff3cd;padding:12px;border-radius:8px;font-size:12px;color:#856404;margin-top:15px}
  .ok{background:#d4edda;padding:15px;border-radius:8px;color:#155724;text-align:center;margin-top:15px}
  .err{background:#f8d7da;padding:15px;border-radius:8px;color:#721c24;margin-top:10px;font-size:14px}
  .loading{text-align:center;padding:20px;color:#666}
  .loading::after{content:'...';animation:dots 1.5s infinite}
  @keyframes dots{0%{content:'.'}33%{content:'..'}66%{content:'...'}}"
  .footer{text-align:center;padding:20px;font-size:12px;color:#999}
</style>
</head>
<body>
<div class="header">🚚 华美物流发货单生成器</div>
<div class="card">
  <form id="form" method=post enctype=multipart/form-data>
    <label>📅 发货日期（如 2026-03-31）</label>
    <input type=text name=ship_date placeholder="2026-03-31" required>

    <label>📁 今日发货数据表（.xlsx）</label>
    <input type=file name=file accept=".xlsx" required>
    <div class=tip>上传"今日发货XXXXXX.xlsx"文件</div>

    <button type=submit id="btn">生成发货单</button>
  </form>
  <div id="msg"></div>
</div>
<div class="footer">📱 可添加到手机主屏幕，像APP一样使用</div>

<script>
if('serviceWorker' in navigator){
  navigator.serviceWorker.register('/sw.js').then(r=>console.log('SW registered'));
}
const form = document.getElementById('form');
const btn = document.getElementById('btn');
const msg = document.getElementById('msg');
form.onsubmit = async e => {
  e.preventDefault();
  btn.disabled = true;
  btn.textContent = '正在生成，请稍候';
  msg.innerHTML = '<div class="loading">正在生成</div>';
  const fd = new FormData(form);
  try {
    const r = await fetch('/', {method:'POST', body:fd});
    if(r.ok){
      const blob = await r.blob();
      const a = document.createElement('a');
      a.href = URL.createObjectURL(blob);
      a.download = decodeURIComponent(r.headers.get('Content-Disposition')?.match(/filename=(.*)/)?.[1] || '华美物流发货单.xlsx');
      a.click();
      msg.innerHTML = '<div class="ok">✅ 生成成功！文件已下载</div>';
    } else {
      const t = await r.text();
      msg.innerHTML = '<div class="err">❌ 错误: '+t+'</div>';
    }
  } catch(err){
    msg.innerHTML = '<div class="err">❌ 网络错误: '+err.message+'</div>';
  }
  btn.disabled = false;
  btn.textContent = '生成发货单';
};
</script>
</body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/manifest.json")
def manifest():
    return Response(
        __import__('json').dumps(MANIFEST),
        mimetype="application/json"
    )

@app.route("/sw.js")
def sw():
    return Response(SW_JS, mimetype="application/javascript")

@app.route("/icon-192.png")
def icon192():
    import base64
    # 生成简单的彩色图标PNG（用纯Python）
    return send_file(generate_icon(192), mimetype="image/png")

@app.route("/icon-512.png")
def icon512():
    return send_file(generate_icon(512), mimetype="image/png")

def generate_icon(size):
    """用纯Python生成带背景色的PNG图标"""
    import zlib, struct, io
    def chunk(tag, data):
        return struct.pack('>I', len(data)) + tag + data
    # 简单纯色图标（蓝色背景+白字）
    def png(w, h, bg_r=68, bg_g=114, bg_b=196):
        # IHDR
        ihdr = struct.pack('>IIBBBBB', w, h, 8, 2, 0, 0, 0)
        # 简单像素：每隔一个像素生成棋盘格
        raw = b''
        for y in range(h):
            raw += b'\x00'  # filter byte
            for x in range(w):
                # 画一个简化的"发"字方块
                cx, cy = x - w//4, y - h//4
                in_box = 0 <= cx < w//2 and 0 <= cy < h//2
                # 简单图案：中心加号
                dist = abs(cx - w//4) + abs(cy - h//4)
                if dist < h//8 and abs(cx - w//4) < w//12:
                    r,g,b = 255,255,255
                elif in_box:
                    r,g,b = 255,220,60
                else:
                    r,g,b = bg_r,bg_g,bg_b
                raw += bytes([r,g,b])
        idat = zlib.compress(raw, 9)
        return io.BytesIO(b'\x89PNG\r\n\x1a\n' + chunk(b'IHDR', ihdr) + chunk(b'IDAT', idat) + chunk(b'IEND', b''))
    return png(size, size)

@app.route("/", methods=["POST"])
def upload():
    f = request.files["file"]
    date = request.form["ship_date"]
    path = os.path.join(UPLOAD, f.filename)
    f.save(path)
    try:
        buf = gen_invoice(path, date)
        date_str = date.replace("-","")
        resp = Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''华美物流发货单{date_str}.xlsx"}
        )
        return resp
    except Exception as e:
        return Response(str(e), status=500)

if __name__ == "__main__":
    print("=" * 50)
    print("  🚚 华美物流发货单生成器 PWA 已启动")
    print("  本地访问: http://localhost:5000")
    print("  手机访问: http://电脑IP:5000")
    print("  按 Ctrl+C 停止")
    print("=" * 50)
    app.run(host="0.0.0.0", port=5000)
